"""KB ingest script: scan kb/sources/, extract text, chunk, embed, store in Chroma.

Usage:
    python scripts/ingest_kb.py [--reset]

Options:
    --reset   Drop existing collection and rebuild from scratch.

Idempotent: re-running updates existing docs by source path.
"""
from __future__ import annotations

import argparse
import asyncio
import hashlib
import logging
import sys
from pathlib import Path

# Make `agent` importable when running from project root
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from agent.tools.kb_index import KBIndex  # noqa: E402

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
# Silence noisy third-party INFO logs (httpx/OpenAI HTTP request lines).
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("openai").setLevel(logging.WARNING)
logging.getLogger("httpcore").setLevel(logging.WARNING)
log = logging.getLogger("ingest_kb")

CHARS_PER_CHUNK = 500
CHUNK_OVERLAP = 50

EXCLUDE_PREFIXES = (".", "_")
TEXT_EXTS = {".md", ".txt", ".rst"}
PDF_EXTS = {".pdf"}
DOCX_EXTS = {".docx"}
XLSX_EXTS = {".xlsx"}


def extract_text(path: Path) -> str:
    """Extract plain text from a file. No prompt-injection wrap (this is trusted source KB)."""
    ext = path.suffix.lower()
    if ext in TEXT_EXTS:
        return path.read_text(encoding="utf-8", errors="ignore")
    if ext in PDF_EXTS:
        import fitz  # pymupdf
        doc = fitz.open(str(path))
        return "\n\n".join(str(p.get_text()) for p in doc)
    if ext in DOCX_EXTS:
        from docx import Document
        doc = Document(str(path))
        return "\n\n".join(p.text for p in doc.paragraphs)
    if ext in XLSX_EXTS:
        from openpyxl import load_workbook
        wb = load_workbook(str(path), read_only=True, data_only=True)
        rows: list[str] = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                rows.append("\t".join(cells))
        return "\n".join(rows)
    raise ValueError(f"Unsupported extension: {ext}")


def chunk_text(text: str, chars: int = CHARS_PER_CHUNK, overlap: int = CHUNK_OVERLAP) -> list[str]:
    """Naive sliding-window chunker on paragraph boundaries, falls back to char window."""
    paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
    chunks: list[str] = []
    buf = ""
    for p in paragraphs:
        if not buf:
            buf = p
            continue
        if len(buf) + len(p) + 2 <= chars:
            buf = buf + "\n\n" + p
        else:
            chunks.append(buf)
            # Keep tail of buf for overlap
            tail = buf[-overlap:] if overlap > 0 and len(buf) > overlap else ""
            buf = tail + "\n\n" + p if tail else p
    if buf:
        chunks.append(buf)
    # If a single paragraph exceeds chars, split hard
    out: list[str] = []
    for c in chunks:
        if len(c) <= chars:
            out.append(c)
            continue
        for i in range(0, len(c), chars - overlap):
            out.append(c[i : i + chars])
    return out


def doc_id_for(source: Path, chunk_index: int) -> str:
    raw = f"{source.as_posix()}::{chunk_index}"
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def list_source_files(root: Path) -> list[Path]:
    files: list[Path] = []
    for p in sorted(root.rglob("*")):
        if not p.is_file():
            continue
        if p.name.startswith(EXCLUDE_PREFIXES):
            continue
        if p.suffix.lower() in TEXT_EXTS | PDF_EXTS | DOCX_EXTS | XLSX_EXTS:
            files.append(p)
    return files


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--reset", action="store_true", help="Drop existing KB and rebuild")
    ap.add_argument("--source", default=None, help="Override kb/sources path")
    args = ap.parse_args()

    from dotenv import load_dotenv
    load_dotenv(Path.cwd() / ".env")

    from agent.storage.paths import get_kb_dir
    kb_dir = Path(args.source) if args.source else get_kb_dir()
    sources = kb_dir / "sources"
    if not sources.exists():
        log.error(f"KB source dir not found: {sources}")
        return 1

    files = list_source_files(sources)
    if not files:
        log.error(f"No source files found in {sources}")
        return 1
    log.info(f"Found {len(files)} source files in {sources}")

    idx = KBIndex.get()
    if args.reset:
        log.info("Resetting KB collection...")
        idx.reset()
        idx.init(kb_dir)

    total_chunks = 0
    skipped: list[str] = []
    for f in files:
        try:
            text = extract_text(f)
        except Exception as e:
            log.warning(f"  skip {f.name}: {e}")
            skipped.append(f.name)
            continue
        chunks = chunk_text(text)
        for i, c in enumerate(chunks):
            # idx.add is async; run each call. Batch add would be faster but
            # batch size constraints differ per embed provider.
            asyncio.run(idx.add(
                doc_id=doc_id_for(f, i),
                text=c,
                metadata={"source": str(f.relative_to(kb_dir)), "chunk_index": i},
            ))
        log.info(f"  + {f.relative_to(kb_dir)}: {len(chunks)} chunks")
        total_chunks += len(chunks)

    log.info(f"Done. Total chunks ingested: {total_chunks}. Skipped: {len(skipped)}.")
    if skipped:
        log.info(f"  Skipped files: {skipped}")
    log.info(f"KB total docs: {idx.count()}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
